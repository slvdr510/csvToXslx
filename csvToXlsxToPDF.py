import csv
import os
import time
from openpyxl import Workbook
from datetime import datetime
import pandas as pd
import pdfkit


##########################################################
#                     MODIFICAR ESTO                     #
nombre_asignatura = "PFIS"
semana_de_docencia = 1
tema_asignatura = "Tema 1 - Introducción"
orden_alfabetico_hacia_abajo = True
##########################################################

##########################################################
#                      MAGIC WORDS                       #
ARCHIVO_CSV = "asistentes.csv"
CABECERA_COLUMNA_NOMBRE = "NOMBRE"
CABECERA_COLUMNA_APELLIDOS = "APELLIDOS"
##########################################################


# Get the absolute path to the current folder
current_folder = os.path.abspath(os.path.dirname(__file__))

fechaActual = datetime.now().strftime("%d-%m-%Y")  # String
fechaActual_formatoIndexable = datetime.now().strftime("%Y-%m-%d")  # String
xlsx_filename = os.path.join(current_folder, f"asistencia_{fechaActual_formatoIndexable}.xlsx")
pdf_filename = os.path.join(current_folder, f"asistencia_{fechaActual_formatoIndexable}.pdf")

# Leer el archivo csv y guardarlo en una lista
nombres_y_apellidos = []
with open(os.path.join(current_folder, ARCHIVO_CSV), newline="", encoding="utf-8") as file:
    reader = csv.reader(file)
    for row in reader:
        nombres_y_apellidos.append(row)

#print(json.dumps(nombres_y_apellidos, indent=4))

# Ordenar los alumnos por apellidos
nombres_y_apellidos = sorted(nombres_y_apellidos, key=lambda x: (x[1], x[0]))

# Creamos el archivo xlsx
workbook = Workbook()

# Seleccionamos la hoja activa
hoja = workbook.active

contador_personas = 1 # Contador para saber cuantos personas

fila = 1 # Fila en la que comienza (El xlsx comienza en 1)
columna = 1 # Columna en la que comienza (El xlsx comienza en 1)

# Iteramos sobre los nombres y apellidos
if orden_alfabetico_hacia_abajo:

    for i in range(len(nombres_y_apellidos)):
        
        hoja.cell(row=fila, column=columna).value = nombres_y_apellidos[i][0] # Nombre
        hoja.cell(row=fila, column=columna+1).value = nombres_y_apellidos[i][1] # Apellido
        
        if contador_personas % 8 == 0:
            fila += 1

        fila += 1
        
        if contador_personas >= len(nombres_y_apellidos)/2:
            if contador_personas % 8 == 0:
                columna = 4
                fila = 1
                contador_personas = 0
                
        contador_personas += 1
        
else:
    for i in range(len(nombres_y_apellidos)):
    
        hoja.cell(row=fila, column=columna).value = nombres_y_apellidos[i][0] # Nombre
        hoja.cell(row=fila, column=columna+1).value = nombres_y_apellidos[i][1] # Apellido
        
        if contador_personas % 16 == 0:
            fila += 1
        
        if contador_personas % 2 == 0:
            fila += 1
        
        if columna % 4 == 0:
            columna = 1
        else:
            columna = 4
            
        contador_personas += 1


# Añadir la cabecera de las columnas Nombre y Apellidos
hoja.insert_rows(1)
hoja.cell(row=1, column=1).value = CABECERA_COLUMNA_NOMBRE
hoja.cell(row=1, column=2).value = CABECERA_COLUMNA_APELLIDOS
hoja.cell(row=1, column=4).value = CABECERA_COLUMNA_NOMBRE
hoja.cell(row=1, column=5).value = CABECERA_COLUMNA_APELLIDOS

# Añadir la cabecera de las columnas: Nombre de la asignatura, Semana de docencia, Tema y Fecha actual
hoja.insert_rows(1)
hoja.insert_rows(1)
hoja.cell(row=1, column=1).value = nombre_asignatura
hoja.cell(row=1, column=2).value = "Semana " + str(semana_de_docencia)
hoja.cell(row=1, column=5).value = tema_asignatura
hoja.cell(row=1, column=4).value = fechaActual

# Set empty cells to an empty string
for row in hoja.iter_rows():
    for cell in row:
        if cell.value is None:
            cell.value = ""

# Set minimum column width for empty columns
for col in hoja.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) if max_length > 0 else 20
    hoja.column_dimensions[column].width = adjusted_width

workbook.save(xlsx_filename)

time.sleep(3)

# Convertir el archivo XLSX a PDF

# Load Excel file
df = pd.read_excel(xlsx_filename)

# Replace "Unnamed" and "NaN" values with empty strings
df.columns = df.columns.str.replace('Unnamed.*', '', regex=True)
df = df.fillna('')

# Convert to HTML
html_string = df.to_html(index=False)

# Add CSS for UTF-8 encoding and proper font display
html_string = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <style>
        table {{
            font-family: Arial, sans-serif;
            border-collapse: collapse;
            width: 100%;
        }}
        th, td {{
            border: 1px solid #dddddd;
            text-align: left;
            padding: 8px;
            min-width: 100px; /* Set minimum width for table cells */
        }}
        th {{
            background-color: #f2f2f2;
        }}
        tr {{
            min-height: 30px; /* Set minimum height for table rows */
        }}
    </style>
</head>
<body>
    {html_string}
</body>
</html>
"""

# Specify the path to wkhtmltopdf
config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
## NOTA: HAY QUE INSTALAR wkhtmltopdf y AÑADIRLO AL PATH DE WINDOWS

# Save to PDF
pdfkit.from_string(html_string, pdf_filename, configuration=config)