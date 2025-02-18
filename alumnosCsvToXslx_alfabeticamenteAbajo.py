import csv
from openpyxl import Workbook
import json

# Leer el archivo csv y guardarlo en una lista
nombres_y_apellidos = []
with open("asistentes.csv", newline="", encoding="utf-8") as file:
    reader = csv.reader(file)
    for row in reader:
        nombres_y_apellidos.append(row)

#print(json.dumps(nombres_y_apellidos, indent=4))

# Ordenar los alumnos por apellidos
nombres_y_apellidos = sorted(nombres_y_apellidos, key=lambda x: (x[1], x[0]))

# Creamos el archivo xlsx
workbook = Workbook()

# Seleccionamos la hoja activa
hoja = workbook.active

contador_personas = 1 # Contador para saber cuantos personas

fila = 1 # Fila en la que comienza (El xlsx comienza en 1)
columna = 1 # Columna en la que comienza (El xlsx comienza en 1)

# Iteramos sobre los nombres y apellidos
for i in range(len(nombres_y_apellidos)):
    
    hoja.cell(row=fila, column=columna).value = nombres_y_apellidos[i][0] # Nombre
    hoja.cell(row=fila, column=columna+1).value = nombres_y_apellidos[i][1] # Apellido
    
    if contador_personas % 8 == 0:
        fila += 1

    fila += 1
    
    if contador_personas >= len(nombres_y_apellidos)/2:
        if contador_personas % 8 == 0:
            columna = 4
            fila = 1
            contador_personas = 0
            
    contador_personas += 1

workbook.save("asistentes.xlsx")